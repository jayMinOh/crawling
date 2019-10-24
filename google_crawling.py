from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import json
import os
from datetime import datetime
import sys, threading, re
from urllib import parse

from PyQt5.QtWidgets import QMainWindow, QApplication, QWidget, QWidget, QPushButton, QDesktopWidget, QFrame, QLabel,QGridLayout, QLineEdit, QCheckBox, QMessageBox, QCalendarWidget, QComboBox, QBoxLayout
from PyQt5.QtGui import QIcon, QImage,QPalette,  QBrush
from PyQt5.QtCore import pyqtSlot, Qt, QSize, QPropertyAnimation, QRect, QDate
from openpyxl import Workbook


class crawlingApp(QMainWindow):
    questions = []

    def __init__(self):
        super().__init__()
        self.init_ui()
 
    def init_ui(self):
        self.main_init()
        self.init_position()
        button = QPushButton('start crawling', self)
        button.setToolTip('start crawling')
        button.move(95, 240)
        
        label = QLabel("검색어 :", self)
        label.move(20, 20)

        self.lineEdit = QLineEdit("", self)
        self.lineEdit.move(60,20)

        label2 = QLabel("기간 :", self)
        label2.move(20, 60)

        self.lineEdit2 = QLineEdit("D/M/YYYY", self)
        self.lineEdit2.resize(100, 30)
        self.lineEdit2.move(60,60)
        
        self.lineEdit3 = QLineEdit("D/M/YYYY", self)
        self.lineEdit3.resize(100, 30)
        self.lineEdit3.move(170,60)

        label3 = QLabel("검색개수 :", self)
        label3.move(20, 100)
        
        self.cb = QComboBox(self)
        self.cb.addItems(["1000", "전체"])
        self.cb.resize(100, 30)
        self.cb.move(80,  100)

        self.frame = QFrame(self)
        self.frame.setGeometry(QRect(20, 140, 200, 90))
        self.frame.setFrameShape(QFrame.Box)
        self.frame.setFrameShadow(QFrame.Raised)

        self.checkBox1 = QCheckBox("twitter", self)
        self.checkBox1.move(30, 140)
        self.checkBox1.resize(150, 30)
        
        self.checkBox2 = QCheckBox("facebook", self)
        self.checkBox2.move(30, 160)
        self.checkBox2.resize(150, 30)

        self.checkBox3 = QCheckBox("instagram", self)
        self.checkBox3.move(30, 180)
        self.checkBox3.resize(150, 30)

        self.checkBox4 = QCheckBox("blog", self)
        self.checkBox4.move(30, 200)
        self.checkBox4.resize(150, 30)

    

        button.clicked.connect(self.startCrawling)
        label.setStyleSheet('QLabel {color : white}')
        label2.setStyleSheet('QLabel {color : white}')
        label3.setStyleSheet('QLabel {color : white}')
        self.setStyleSheet('QCheckBox {color : white}')
        self.show()

    def main_init(self) :
        self.setGeometry(100, 100, 280, 280)  # Setting App Size
        self.setWindowTitle('Crawling')     # Setting App Title
        self.setWindowIcon(QIcon('./images/favicon.png'))    # Setting App favicon
        # 배경 색상 넣기.
        # self.setAutoFillBackground(True)
        # p = self.palette()
        # p.setColor(self.backgroundRole(), Qt.blue)
        # self.setPalette(p)
        oImage = QImage("./images/background_image.jpg")
        sImage = oImage.scaled(QSize(300, 200))
        palette = QPalette()
        palette.setBrush(10, QBrush(sImage))
        self.setPalette(palette)
    
    def init_position(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    
    def startCrawling(self) :
      searchList = self.makeList()
      searchWord = self.lineEdit.text()
      startDate = self.lineEdit2.text()
      endDate = self.lineEdit3.text()
      if startDate == "D/M/YYYY" :
        startDate = ''

      if endDate == "D/M/YYYY" :
        endDate = ''
      
      if searchWord == '':
        QMessageBox.information(self, '경고', '검색어를 입력 해주세요.')
        return 0

      self.startDate = startDate
      self.endDate = endDate
      self.searchWord = searchWord
      for i in searchList :       
        t = threading.Thread(target=crawlingApp.crawling, args=(self, i))
        t.start()

    def makeList(self) :
      searchList = []
      twitterCheck = self.checkBox1.isChecked()
      facebookCheck = self.checkBox2.isChecked()
      instagramCheck = self.checkBox3.isChecked()
      blogCheck = self.checkBox4.isChecked()

      if twitterCheck:
        searchList.append("twitter")
      if facebookCheck:
        searchList.append("facebook")
      if instagramCheck:
        searchList.append("instagram")
      if blogCheck:
        searchList.append("blog")
      
      return searchList

    def crawling(self, key) :
      
      path = os.getcwd()+ "/chromedriver"
      driver = webdriver.Chrome(path)
      cnt = int(self.cb.currentText())
      url = "https://www.google.co.kr/search?q=" + key+":"+ self.searchWord + "&start=pageNo"
      
      if self.startDate != '':
       url += "&tbs=cdr:1,cd_min:" + self.startDate  + ",cd_max:" + self.endDate
      
      contents = []
      driver.get(url.replace("pageNo" , str(0)))

      if cnt != 1000:
        driver.find_element(By.XPATH,'//*[@id="hdtb-tls"]').click()
        time.sleep(1)
        resultCnt = driver.find_element(By.ID, 'resultStats').text
        searchStartStr = "약"
        searchEndStr = "개"
        startIndex = resultCnt.find(searchStartStr)
        endIndex = resultCnt.find(searchEndStr)
        resultCnt = int(resultCnt[int(startIndex):int(endIndex+1)].replace("약","").replace("개","").replace(" ","").replace(",",""))
        loopCnt = int(resultCnt/10) + 1 
      else :
        loopCnt = int(cnt/10) + 1

      for k in range(loopCnt) :
        driver.get(url.replace("pageNo" , str(k * 10)))
        time.sleep(1)
        pattern = re.compile(r'\s+')
        loopFlag = re.sub(pattern, '',driver.find_element(By.ID, 'topstuff').text) != "" # 검색결과 체크
        if loopFlag == True:  # 검색결과가 없을 경우
          break
        
        else : # 검색결과가 있을 경우
          for j in range(10):
            urlXpath = '//*[@id="rso"]/div/div/div[idx]/div/div/div[1]/a'
            titleXpath = '//*[@id="rso"]/div/div/div[idx]/div/div/div[1]/a/h3'
            contentXpath = '//*[@id="rso"]/div/div/div[idx]/div/div/div[2]/div/span'
            urlXpath = urlXpath.replace("idx", str(j+1))
            titleXpath = titleXpath.replace("idx", str(j+1))
            contentXpath = contentXpath.replace("idx", str(j+1))
            try:  #검색결과 내에서 크롤링중 데이터가 있을 경우
              contents_url = driver.find_element(By.XPATH, urlXpath).get_attribute("href")
              contents_title = driver.find_element(By.XPATH, titleXpath).text
              contents_content = driver.find_element(By.XPATH, contentXpath).text
              obj = {
                "url" : contents_url,
                "title" : contents_title, 
                "content" : contents_content
              }
              contents.append(json.dumps(obj, ensure_ascii=False)) # json으로 변환 시에, 한글이 깨지는 문제가 발생함..
            except: # 검색 결과 내에서 크롤링 데이터가 없을 경우
              break

      self.makeExcel(contents, key)
      driver.close()  
      return 1

    def makeExcel(self, contents, key) :
      write_wb = Workbook()
      write_ws = write_wb.active
      write_ws['A1'] = "URL"

      write_ws['B1'] = "제목"
      write_ws['C1'] = "내용"

      i = 1
      for content in contents:
        i = i+1 
        content = json.loads(content)
        write_ws.cell(i,1, content["url"])
        write_ws.cell(i,2, content["title"])
        write_ws.cell(i,3, content["content"])
      fileName = os.getcwd() +"/excel/" + key + "_" +  self.searchWord + "_" + 'excel.xlsx'
      write_wb.save(fileName)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = crawlingApp()
    sys.exit(app.exec_())